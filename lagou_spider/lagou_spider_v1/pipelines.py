# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

from lagou_spider_v1.items import Position_Url_List
from lagou_spider_v1.items import Position_Info
from openpyxl import Workbook,load_workbook

import pymongo

class PositionPipeline(object):
    
    def save_position_data_to_excel(self,items):

        # 程序流程：每次从表格中读取内容，读取行数，然后写入最后一行，第一次读取时候，创建好表格
        filepath = r"date.xlsx"
        try:
            # 打开表格文件
            position_wb = load_workbook(filename=filepath)
            pass
        except Exception as e:
          
            # 出错就创建新的文件
            wb = Workbook()
            ws1 = wb.active        
            # 表头
            ws1["A1"] = r"岗位"
            ws1["B1"] = r"地区"        
            ws1["C1"] = r"公司"
            ws1["D1"] = r"薪资"
            ws1["E1"] = r"详情"
            ws1["F1"] = r"链接"
            # 保存到本地
            wb.save(filepath)
            print("初始化创建成功，已保存本地")

            position_wb = load_workbook(filename=filepath)

        # 读取行数，将数据写入最后一行
        position_ws1 = position_wb.active # 激活表1
        line_number = 1
        for i in position_ws1.rows:
            line_number = line_number + 1
        print("-------行高---------")
        print(line_number)  
        # 在最后一行写入item数据
        position_ws1.cell(row=line_number,column=1).value = items["position_name"]
        position_ws1.cell(row=line_number,column=2).value = items["position_location"]
        position_ws1.cell(row=line_number,column=3).value = items["position_company"]
        position_ws1.cell(row=line_number,column=4).value = items["position_pay"]
        position_ws1.cell(row=line_number,column=5).value = items["position_detail"]
        position_ws1.cell(row=line_number,column=6).value = items["position_url"]
        
        # 保存到表格
        position_wb.save(filename=filepath)

        return True

        # 每一个item都是独立的数据


    def process_item(self, item, spider):
        
        # 如果是职位url处理，就保存本地
        if isinstance(item,Position_Url_List):

            print "----------------到这里了！！！88888888---------------"
            with open("position_url.txt","a") as f :
                for url in item["position_url"]:
                    f.write(url  + "\n")
        if isinstance(item,Position_Info):

            print "到这里了！！！！！！！！！！！！！！！！！0000000000000"
            self.save_position_data_to_excel(item)


        # 如果是职位详情内容处理，也保存本地
        return item


class DbPipeline(object):
    # 把数据保存到mongo数据库中
    # 需要功能：1、初始化数据库。 2、插入功能 3、查询功能

    def __init__(self):
        # 初始化流程：链接数据库，获取数据库db对象，
        self.client = pymongo.MongoClient('localhost',27017)
        self.db = self.client['lagou']
        print("数据库初始化成功")
    

    def insert_to_db(self,item):

        if isinstance(item,Position_Url_List):
            # 处理逻辑，url放到url集合中，position数据放到position数据集合中
            for url in item['position_url']:
                self.db.url_3.insert_one({'url':str(url)})

            self.client.close()
            print("插入数据成功！！！")
            return True


        if isinstance(item,Position_Info):
            # 保存到position中
            # 构建字典
            dict = {
            'position_name':item['position_name'],
            'position_company':item['position_company'],
            'position_pay':item['position_pay'],
            'position_location': item['position_location'],
            'position_detail':item['position_detail'],
            'position_url':item['position_url']
            }
            self.db.position_3.insert_one(dict)
            self.client.close()
            pass


    def process_item(self,item,spider):
        # 写入数据
        print self.insert_to_db(item)
        return item



